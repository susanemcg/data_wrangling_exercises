# An example of reading data from a .xlsx file with Python, using the "openpyxl" library.
# First, you'll need pip install the openpyxl library: https://pypi.org/project/openpyxl/

# specify the particular "chapter" you want to import the "openpyxl" library
# in this case, "load_workbook"
from openpyxl import load_workbook

# we'll also import the "csv" library because we want to convert our workbook
# to a `.csv`
import csv

# because this is a very specialized library, there are fewer functions that do more in one step
# pass our source filename as an ingredient to the the openpyxl library's load_workbook "recipe"
# and store the result in a variable called `source_workbook`
source_workbook = load_workbook(filename = 'fredgraph.xlsx')

# an .xlsx workbook can have multiple sheets
# like the "DictReader" function, load_workbook includes useful information,
# like a list that shows us the names of all the data sheets in our workbook
print(source_workbook.sheetnames)

# even though our example workbook only includes one worksheet
# we might have more in the future. So we'll use the "enumerate" functions
# to get both an iterator *and* the sheet name. This will help us
# create one `.csv`file per worksheet
for sheet_num, sheet_name in enumerate(source_workbook.sheetnames):
    # we'll create a variable that points to the current worksheet by
    # passing the current value of `sheet_name` to `source_workbook`
    current_sheet = source_workbook[sheet_name]
    # let's print the value of `sheet_name` just to see what that value is
    print(sheet_name)

    # for each sheet in our workbook, we'll create a separate `.csv` file
    # to do this, we "open" a new file, but make it *writable* ("w")
    # instead of *readable* ("r"), as we have in previous examples
    # for now, we'll just name it "xlsx_"+sheet_name
    output_file = open("xlsx_"+sheet_name+".csv","w")

    # there is a "writer" recipe that lets us easily write `.csv`-formatted rows
    # so, just as we did when "reading", now that we've opened our `output_file`
    # we'll use this recipe to easily write rows, instead of reading them
    output_writer = csv.writer(output_file)

    # now, we need to loop through every row in our sheet
    # the function `iter_rows()` is specific to the `openpyxl` library and
    # converts the rows of `source_workbook` into a list that can be *iterated*, or looped, over
    for row in current_sheet.iter_rows():
        # because of the way that the `openpyxl` library works, if we try to just print the
        # rows directly, we'll get sort of unhelpful cell locations, rather than the data values
        # they contain. So we need to make *another* loop, to go through every cell in every row
        # one at a time. We'll print both the cell location and the value here, though only the latter
        # will be actually written to our new file

        # we'll create an empty list where we'll put the actual values of the cells in each row
        row_cells = []

        # for every cell (or column) in each row....
        for cell in row:

            # let's print what's in here, just to see how the code sees it
            print(cell, cell.value)

            # just add the values to our list, so we get a nice clean `.csv`
            # `append` is a method/recipe that we can use on lists to add things to the end
            row_cells.append(cell.value)


        # notice that the code below is left-aligned with the `for cell in row` code above
        # this means that it will only be run *after* all the cells in a given row have been
        # gone through, with its values appended to our list
        # now we'll actually write these rows to the output file
        output_writer.writerow(row_cells)

    # just for good measure, let's close the `.csv` file we just wrote all that
    # data to...
    output_file.close()
