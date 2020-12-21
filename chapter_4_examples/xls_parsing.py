# A simple example of reading data from a .xlsx file with Python, using the "openpyxl" library.



# first, pip install the openpyxl library: https://pypi.org/project/openpyxl/
from openpyxl import load_workbook


wb = load_workbook(filename = 'fredgraph.xls')


for sheet_name in wb.sheetnames:
    current_sheet = wb[sheet_name]
    for row in current_sheet.iter_rows():
        for cell in row:
            print(cell, cell.value)
